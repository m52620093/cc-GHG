from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import column_index_from_string

from .config import SourceConfig

REQUIRED_COLUMN_KEYS = ("部材図面", "GHG排出量算定パターン", "登録日時", "GHG排出量単位毎")


@dataclass
class SourceRow:
    part_drawing: str
    pattern: str
    registered_at: Any
    ghg_per_unit: float | None


def validate_column_range(config: SourceConfig) -> list[str]:
    """マッピングされた列が読み取り範囲(start_column〜end_column)の外にある場合、警告メッセージを返す。"""
    start = column_index_from_string(config.start_column)
    end = column_index_from_string(config.end_column)
    warnings: list[str] = []
    for key, column in config.columns.items():
        idx = column_index_from_string(column)
        if not (start <= idx <= end):
            warnings.append(
                f"列 '{key}' ({column}) が読み取り範囲 {config.start_column}〜{config.end_column} の外にあります"
            )
    return warnings


def _to_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def read_source_rows(file_path: Path, config: SourceConfig) -> list[SourceRow]:
    for key in REQUIRED_COLUMN_KEYS:
        if key not in config.columns:
            raise ValueError(f"setting.json の source.columns に '{key}' が設定されていません")

    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb[config.sheet_name]
    columns = config.columns
    key_col = columns["部材図面"]

    rows: list[SourceRow] = []
    row_idx = config.data_start_row
    while True:
        key_value = ws[f"{key_col}{row_idx}"].value
        if key_value is None or str(key_value).strip() == "":
            break

        ghg_raw = ws[f"{columns['GHG排出量単位毎']}{row_idx}"].value
        ghg_value = float(ghg_raw) if ghg_raw not in (None, "") else None

        rows.append(
            SourceRow(
                part_drawing=_to_text(key_value),
                pattern=str(ws[f"{columns['GHG排出量算定パターン']}{row_idx}"].value or ""),
                registered_at=ws[f"{columns['登録日時']}{row_idx}"].value,
                ghg_per_unit=ghg_value,
            )
        )
        row_idx += 1

    return rows
