from __future__ import annotations

from datetime import datetime
from pathlib import Path

import openpyxl

from .config import ImportConfig
from .selector import SelectedRow

REQUIRED_MAPPING_KEYS = ("部材図面", "GHG排出量単位毎")


def resolve_output_file_name(pattern: str, now: datetime | None = None) -> str:
    now = now or datetime.now()
    return pattern.replace("{yyyyMMdd}", now.strftime("%Y%m%d"))


def write_import_file(template_path: Path, output_path: Path, config: ImportConfig, rows: list[SelectedRow]) -> None:
    for key in REQUIRED_MAPPING_KEYS:
        if key not in config.column_mapping:
            raise ValueError(f"setting.json の import.column_mapping に '{key}' が設定されていません")

    wb = openpyxl.load_workbook(template_path)
    ws = wb[config.sheet_name]
    mapping = config.column_mapping

    for offset, row in enumerate(rows):
        row_idx = config.data_start_row + offset
        ws[f"{mapping['部材図面']}{row_idx}"] = row.part_drawing
        ws[f"{mapping['GHG排出量単位毎']}{row_idx}"] = row.ghg_per_unit * config.ghg_per_unit_multiplier
        for column, value in config.fixed_values.items():
            ws[f"{column}{row_idx}"] = value

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
